"""Export batch detour results to CSV and XLSX for analysis in Excel.

    nzcl-export --snapshot <id>

Excel is the downstream reporting surface, not the map. The workbook is written
as styled ranges rather than Excel Table objects, and every value is a number
computed by the validated backend - there are no formulas that could drift from
the engine.

Sheets:
  Link Detours       one row per link and direction
  Network Metadata   snapshot provenance
  Quality Summary    QA counts and issues
  Metric Definitions what each column means, and what it does not mean
  Source Lineage     where every input came from
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db
from .config import get_settings

COLUMNS: list[tuple[str, str, int]] = [
    ("snapshot_id", "Snapshot ID", 34),
    ("amds_id", "AMDS Link ID", 40),
    ("link_id", "Internal Link ID", 14),
    ("closure_group_id", "Closure Group ID", 40),
    ("road_name", "Road Name", 30),
    ("rca_name", "Controlling Authority", 26),
    ("road_class", "Road Class", 14),
    ("lifeline", "Lifeline Route", 13),
    ("direction", "Direction", 10),
    ("closure_scope", "Closure Scope", 13),
    ("vehicle_profile", "Vehicle Profile", 14),
    ("status", "Status", 22),
    ("selected_link_length_m", "Selected Link Length (m)", 20),
    ("normal_path_distance_m", "Normal Path Distance (m)", 21),
    ("alternative_distance_m", "Alternative Distance (m)", 21),
    ("added_distance_vs_link_m", "Added Distance vs Link (m)", 23),
    ("network_penalty_m", "Network Penalty (m)", 18),
    ("detour_ratio_vs_link", "Detour Ratio vs Link", 18),
    ("normal_path_time_s", "Normal Time est. (s)", 18),
    ("alternative_time_s", "Alternative Time est. (s)", 21),
    ("added_time_s", "Added Time est. (s)", 18),
    ("disconnected", "No Detour Exists", 15),
    ("corridor_status", "Corridor Status", 15),
    ("corridor_penalty_m", "Corridor Penalty (m)", 19),
    ("isolation_side", "Stranded Side", 14),
    ("isolation_link_count", "Stranded Links", 14),
    ("isolation_length_m", "Stranded Road Length (m)", 21),
    ("speed_source", "Speed Source", 22),
    ("quality_flags", "Quality Flags", 60),
    ("calculated_at_utc", "Calculated At (UTC)", 22),
    ("permalink", "Open in App", 40),
]

DEFINITIONS: list[tuple[str, str]] = [
    ("Selected Link Length (m)",
     "Length of the closed link, computed from its polyline in EPSG:2193."),
    ("Normal Path Distance (m)",
     "Shortest distance between the closed link's own endpoints on the INTACT "
     "network. Often shorter than the link itself when a shortcut exists."),
    ("Alternative Distance (m)",
     "Shortest distance between the same endpoints after every link in the "
     "closure group is removed. This is the replacement path."),
    ("Added Distance vs Link (m)",
     "Alternative Distance minus Selected Link Length. Negative when the closed "
     "link was not itself the shortest way between its endpoints."),
    ("Network Penalty (m)",
     "Alternative Distance minus Normal Path Distance. The more rigorous "
     "measure: it does not assume the closed link was the normal route."),
    ("Detour Ratio vs Link",
     "Alternative Distance divided by Selected Link Length. A ratio of 3 means "
     "the replacement path is three times the length of the closed road."),
    ("Status",
     "OK = a replacement path exists. DISCONNECTED = none exists between the "
     "link's endpoints. UNRESOLVED_TIMEOUT = the search ran out of budget and "
     "the answer is UNKNOWN, not 'no detour'. Other values are application "
     "faults, never network findings."),
    ("No Detour Exists",
     "Yes when Status is DISCONNECTED. On a one-way carriageway this is routine "
     "and does NOT mean the area is cut off - read Corridor Status and "
     "Stranded Links."),
    ("Corridor Status / Penalty",
     "Through-trip comparison between the nearest upstream and downstream points "
     "at which a driver has a choice. Usually the meaningful number for one-way "
     "carriageways."),
    ("Stranded Side / Links / Road Length",
     "What is cut off when no replacement path exists. A handful of links is a "
     "cul-de-sac; hundreds is a community with a single road in."),
    ("Normal / Alternative / Added Time (s)",
     "ESTIMATED travel times. AMDS publishes no speed attribute; speeds are "
     "inferred from urban/rural classification where available, otherwise asset "
     "type and ownership. Never treat these as observed or posted travel times."),
    ("Quality Flags",
     "Machine-readable caveats attached to the result. See "
     "docs/METRIC_DEFINITIONS.md."),
    ("IMPORTANT",
     "These are STRUCTURAL replacement paths. They do not predict how much "
     "traffic uses each alternative route. That requires an origin-destination "
     "demand matrix, capacities, congestion functions and a traffic-assignment "
     "model, none of which is present."),
]

LINEAGE: list[tuple[str, str, str]] = [
    ("AMDS Network Model layer 1",
     "services.arcgis.com/CXBb7LAjgIIdcsPt .../AMDS_NetworkModel_PROD/FeatureServer/1",
     "Link geometry, direction of travel, mode permissions, ownership"),
    ("AMDS RouteName + join table", "Same service, tables 11 and 13",
     "Road names and state-highway numbers"),
    ("AMDS UrbanRural", "Same service, table 12",
     "Urban/rural classification driving the speed estimate"),
    ("AMDS RestrictedTurn", "Same service, table 9",
     "Banned manoeuvres (only 60 exist nationally)"),
    ("AMDS Restriction", "Same service, table 10",
     "Height and weight limits, recorded as flags only"),
    ("AMDS Authority", "Same service, table 2", "Controlling authority names"),
    ("LINZ Basemaps", "basemaps.linz.govt.nz",
     "Web map background only - not the routing network"),
    ("NOT USED", "OpenStreetMap",
     "No OSM data is present. Any future enrichment must stay separable for "
     "ODbL reasons."),
    ("NOT USED", "National Speed Limit Register",
     "Not integrated. Would replace estimated speeds and set speed source to nslr."),
    ("NOT USED", "Traffic counts / AADT",
     "Not integrated. Needed before any statement about vehicles affected."),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_ERR_FILL = PatternFill("solid", fgColor="FFD6D6")


def _rows(snapshot_id: str, base_url: str) -> list[dict[str, Any]]:
    return db.query(
        """
        SELECT d.snapshot_id, d.link_id, l.amds_id, d.closure_group_id,
               l.road_name, l.rca_name, l.model_asset_type, l.lifeline_route,
               d.direction, d.closure_scope, d.vehicle_profile, d.status,
               d.selected_link_length_m, d.normal_path_distance_m,
               d.alternative_distance_m, d.added_distance_vs_link_m,
               d.network_penalty_m, d.detour_ratio_vs_link,
               d.normal_path_time_s, d.alternative_time_s, d.added_time_s,
               d.corridor_status, d.corridor_penalty_m,
               d.isolation_side, d.isolation_link_count, d.isolation_length_m,
               l.speed_source, d.quality_flags, d.calculated_at_utc, d.metric
        FROM detour_results d
        JOIN links l ON l.snapshot_id = d.snapshot_id AND l.link_id = d.link_id
        WHERE d.snapshot_id = %s
        ORDER BY d.link_id, d.direction
        """,
        (snapshot_id,),
    )


def _flatten(r: dict[str, Any], base_url: str) -> dict[str, Any]:
    out = dict(r)
    out["road_class"] = "Roadway" if r["model_asset_type"] == 1 else \
        f"type {r['model_asset_type']}"
    out["lifeline"] = "Yes" if r["lifeline_route"] else "No"
    out["disconnected"] = "Yes" if r["status"] == "DISCONNECTED" else "No"
    out["quality_flags"] = " ".join(r["quality_flags"] or [])
    out["calculated_at_utc"] = r["calculated_at_utc"].isoformat()
    out["permalink"] = (
        f"{base_url}/?link={r['amds_id']}&snapshot={r['snapshot_id']}"
        f"&metric={r['metric']}&vehicle={r['vehicle_profile']}"
        f"&scope={r['closure_scope']}&direction={r['direction']}"
    )
    for k in ("selected_link_length_m", "normal_path_distance_m",
              "alternative_distance_m", "added_distance_vs_link_m",
              "network_penalty_m", "normal_path_time_s", "alternative_time_s",
              "added_time_s", "corridor_penalty_m", "isolation_length_m"):
        if out.get(k) is not None:
            out[k] = round(out[k], 1)
    if out.get("detour_ratio_vs_link") is not None:
        out["detour_ratio_vs_link"] = round(out["detour_ratio_vs_link"], 3)
    return out


def run(snapshot_id: str, out_dir: Path | None = None) -> tuple[Path, Path]:
    s = get_settings()
    out_dir = out_dir or (s.data_dir / "exports")
    out_dir.mkdir(parents=True, exist_ok=True)

    raw = _rows(snapshot_id, s.application_base_url)
    if not raw:
        raise SystemExit(
            f"no detour results for {snapshot_id}; run nzcl-detours first")
    rows = [_flatten(r, s.application_base_url) for r in raw]
    print(f"  {len(rows)} rows (one per link and direction)")

    stem = f"{snapshot_id}-detours"
    csv_path = out_dir / f"{stem}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([h for _, h, _ in COLUMNS])
        for r in rows:
            w.writerow([r.get(k) for k, _, _ in COLUMNS])
    print(f"  wrote {csv_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Link Detours"
    ws.append([h for _, h, _ in COLUMNS])
    for i, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    hdr = ws[1]
    for c in hdr:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    status_col = [k for k, _, _ in COLUMNS].index("status") + 1
    for r in rows:
        ws.append([r.get(k) for k, _, _ in COLUMNS])
    for i in range(2, len(rows) + 2):
        cell = ws.cell(row=i, column=status_col)
        if cell.value == "DISCONNECTED":
            cell.fill = _WARN_FILL
        elif cell.value != "OK":
            cell.fill = _ERR_FILL
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    meta = db.query_one(
        "SELECT * FROM network_snapshots WHERE snapshot_id=%s", (snapshot_id,))
    counts = db.query_one(
        "SELECT (SELECT count(*) FROM links WHERE snapshot_id=%(s)s) AS links, "
        "(SELECT count(*) FROM arcs WHERE snapshot_id=%(s)s) AS arcs, "
        "(SELECT count(*) FROM nodes WHERE snapshot_id=%(s)s) AS nodes, "
        "(SELECT count(*) FROM turn_restrictions WHERE snapshot_id=%(s)s) AS turns",
        {"s": snapshot_id})

    m = wb.create_sheet("Network Metadata")
    m.column_dimensions["A"].width = 34
    m.column_dimensions["B"].width = 96
    for k, v in [
        ("Snapshot ID", meta["snapshot_id"]),
        ("Source dataset", meta["source_dataset"]),
        ("Source URL", meta["source_url"]),
        ("Retrieved (UTC)", meta["retrieved_at_utc"].isoformat()),
        ("Snapshot status", meta["status"]),
        ("Filter applied", meta["where_clause"]),
        ("Raw SHA-256", meta["raw_sha256"]),
        ("Processing version", meta["processing_version"]),
        ("Service feature count", meta["source_feature_count"]),
        ("Downloaded features", meta["downloaded_feature_count"]),
        ("Graph links (after splitting)", counts["links"]),
        ("Graph arcs", counts["arcs"]),
        ("Graph nodes", counts["nodes"]),
        ("Turn restrictions applied", counts["turns"]),
        ("Licence", meta["licence"]),
        ("Attribution", meta["attribution"]),
    ]:
        m.append([k, str(v)])
        m.cell(row=m.max_row, column=1).font = Font(bold=True)
    m.append([])
    m.append(["Ingest notes"])
    m.cell(row=m.max_row, column=1).font = Font(bold=True)
    for n in meta["notes"] or []:
        m.append(["", n])

    q = wb.create_sheet("Quality Summary")
    q.column_dimensions["A"].width = 40
    q.column_dimensions["B"].width = 100
    status_counts: dict[str, int] = {}
    flag_counts: dict[str, int] = {}
    for r in rows:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1
        for f in (r["quality_flags"] or "").split():
            flag_counts[f] = flag_counts.get(f, 0) + 1
    q.append(["Result status counts"])
    q.cell(row=q.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(status_counts.items(), key=lambda x: -x[1]):
        q.append([k, v])
    q.append([])
    q.append(["Quality flag counts"])
    q.cell(row=q.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(flag_counts.items(), key=lambda x: -x[1]):
        q.append([k, v])
    q.append([])
    q.append(["Source-data QA issues"])
    q.cell(row=q.max_row, column=1).font = Font(bold=True)
    for i in db.query(
        "SELECT severity, issue_type, count, detail FROM qa_issues "
        "WHERE snapshot_id=%s ORDER BY count DESC", (snapshot_id,)
    ):
        q.append([f"[{i['severity']}] {i['issue_type']} ({i['count']})", i["detail"]])

    d = wb.create_sheet("Metric Definitions")
    d.column_dimensions["A"].width = 34
    d.column_dimensions["B"].width = 120
    d.append(["Column", "Definition"])
    for c in d[1]:
        c.font = Font(bold=True)
    for k, v in DEFINITIONS:
        d.append([k, v])
        d.cell(row=d.max_row, column=2).alignment = Alignment(wrap_text=True)

    ln = wb.create_sheet("Source Lineage")
    for col, width in (("A", 30), ("B", 62), ("C", 62)):
        ln.column_dimensions[col].width = width
    ln.append(["Input", "Source", "Used for"])
    for c in ln[1]:
        c.font = Font(bold=True)
    for row in LINEAGE:
        ln.append(list(row))

    xlsx_path = out_dir / f"{stem}.xlsx"
    wb.save(xlsx_path)
    print(f"  wrote {xlsx_path}")
    print(f"\n  status: {status_counts}")
    return csv_path, xlsx_path


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Export detour results")
    ap.add_argument("--snapshot")
    ap.add_argument("--out-dir", type=Path)
    args = ap.parse_args(argv)

    snap = args.snapshot
    if not snap:
        row = db.query_one("SELECT snapshot_id FROM network_snapshots "
                           "ORDER BY retrieved_at_utc DESC LIMIT 1")
        if not row:
            raise SystemExit("no snapshots")
        snap = row["snapshot_id"]
    print(f"exporting {snap}")
    run(snap, args.out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
